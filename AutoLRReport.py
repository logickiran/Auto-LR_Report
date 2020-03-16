import xlrd
import xlwt
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles.fills import PatternFill



#Input files
inputfile = 'Raw_Results.xlsx'
outputfile = "Siebel_GUI_Regression_Report_Feb'20_old.xlsx"

print(f'=======Input File ============ {inputfile}')
print(f'=======Regression Input File ============ {outputfile}')


#Ignore list of account names
ignore_list = ['NFR_Siebel_GUI_Regression_015-SYM_FX_Adoption_Wechsel_zu_AllIP','NFR_Siebel_GUI_Regression_016-SYM_FX_Adoption_Validate_Quote','NFR_Siebel_GUI_Regression_017-SYM_FX_Adoption_Submit_Quote']

#defining borders
thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
thick_border = Border(left=Side(style='medium'),
                         right=Side(style='medium'),
                         top=Side(style='medium'),
                         bottom=Side(style='medium'))


#Defining date and variables
currentDay = datetime.now().day
currentMonth = datetime.now().month
currentYear = str(datetime.now().year)
month = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October",
             "November", "December"]
currentMonth_word = month[currentMonth]
current_date = str(currentDay) + '.' + str(currentMonth) + '.' + currentYear


def apply_border_and_alignment(cell):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def check_value(insert_row, insert_value):
    #loading output file.
    wb = openpyxl.load_workbook(outputfile)
    sheet = wb.worksheets[0]


    # Checking if the current date is already there in 3 column(checking only third column)
    # If there and updating value
    # else creating new column
    if current_date in sheet.cell(row=4, column=3).value:
        sheet.cell(row=insert_row, column=3, value=insert_value)
        cell = 'C' + str(insert_row)
        apply_border_and_alignment(sheet[cell])

        try:
            thrashhold_value = float(sheet.cell(row=insert_row, column=2).value)

            #comparing thrashhold value with average value
            if insert_value > thrashhold_value:
                sheet[cell].fill = PatternFill(start_color="ff0000", fill_type="solid")
            else:
                sheet[cell].fill = PatternFill(start_color="92d050", fill_type="solid")
        except:
            sheet[cell] = 'N/A'

    else:
        sheet.insert_cols(3)
        sheet.column_dimensions['C'].width = 30
        #First row
        first_row = sheet.cell(row=1, column=3, value=currentMonth_word + "'" + currentYear + ' MONTHLY RELEASE')
        first_row.font = first_row.font.copy(bold=True, size="10")
        apply_border_and_alignment(sheet['C1'])
        sheet['C1'].border = thick_border

        #third row
        apply_border_and_alignment(sheet['C3'])
        sheet['C3'].border = thick_border
        sheet['C3'].fill = PatternFill(start_color="d9d9d9", fill_type="solid")

        #fourth row
        sheet.cell(row=4, column=3,
                   value='Symphony Regression \n(' + str(currentDay) + '.' + str(currentMonth) + '.' + currentYear + ')')
        sheet['C4'].fill = PatternFill(start_color="d9d9d9", fill_type="solid")
        sheet['C4'].font = Font(bold=True,size = "10")
        apply_border_and_alignment(sheet['C4'])
        sheet['C4'].border = thick_border

        #fifth row
        sheet['C5'].fill = PatternFill(start_color="003366", fill_type="solid")

        sheet.cell(row=insert_row, column=3, value=insert_value)

    wb.save(outputfile)




# Creating Input file sheet
wb_input = xlrd.open_workbook(inputfile)
sheet_input = wb_input.sheet_by_index(0)

#Creating output file sheet
wb_output = xlrd.open_workbook(outputfile)
sheet_out = wb_output.sheet_by_index(0)


# wb2 = xlwt.Workbook(outputfile)
# sheet = wb2.get_active_sheet



#Get all names from output files in a list to compare with names of raw input file
transaction_name_output_file = []


#Iterating through all template file names and adding in above list.
for i in range(sheet_out.nrows):
    transaction_name_output_file.append(sheet_out.cell_value(i, 0))

# Finding no of rows and columns in input sheets
rows, cols = sheet_input.nrows, sheet_input.ncols
print('=======Starting putting values========')
#Going through every row of input record.
for row in range(0, rows):
    raw_sheet_rows = sheet_input.row_values(row)

    if raw_sheet_rows[1] == '' or raw_sheet_rows[1] in ignore_list:
        continue
    if raw_sheet_rows[1] in transaction_name_output_file:
        insert_row = transaction_name_output_file.index(raw_sheet_rows[1]) + 1
        insert_value = round(float(raw_sheet_rows[4]), 2)
        check_value(insert_row, insert_value)




# Handling null fields with N/A
wb = openpyxl.load_workbook(outputfile)
sheet_out_new = wb.worksheets[0]
rows, cols = sheet_out_new.max_row, sheet_out_new.max_column

for row in range(6, rows+1):
    cell = 'C' + str(row)
    if sheet_out_new.cell(row=row, column=1).value in ignore_list:
        continue
    if sheet_out_new.cell(row=row,column=3).value == None:
        try:
            sheet_out_new.cell(row=row, column=3, value='N/A')
            sheet_out_new[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet_out_new[cell].border = thin_border
        except:
            pass


#Formatting file name

outputfile_name = 'Siebel_GUI_Regression_Report' + '_' + currentMonth_word[:3] + "'" + str(currentYear)[2:] + '.xlsx'
wb.save(outputfile_name)
print('=======Done===========================')
